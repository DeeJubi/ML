# My Logger
import  openpyxl, os, datetime


home = os.path.expanduser('~')
print( 'Starting process...')

#load excel sheet, delete columns and save
wb1 = openpyxl.load_workbook(home + '\\Desktop\\Book1.xlsx')
sheet1 = wb1.active
y = 1
count = 0

print("Enter date (e.g DD/MM/YYYY): ")
date = input()

mr = sheet1.max_row + 1


while y == 1:

    #Date
    sheet1.cell(row = mr, column = 1).value = date

    #Project task
    proj_task = '''
    1. READ
    2. BUILD
    3. MEETING
    4. RESEARCH
    5. DOCUMENTATION
    6. PREPARATION
    '''
    print("Choose PROJECT TYPE: ")
    print(proj_task)
    x = input()
    if x == '1':
       sheet1.cell(row = mr, column = 2).value = "READ" 
    elif x == '2' :
       sheet1.cell(row = mr, column = 2).value = "BUILD"
    elif x == '3' :
       sheet1.cell(row = mr, column = 2).value = "MEETING"
    elif x == '4' :
       sheet1.cell(row = mr, column = 2).value = "RESEARCH"
    elif x == '5' :
       sheet1.cell(row = mr, column = 2).value = "DOCUMENTATION"
    elif x == '6' :
       sheet1.cell(row = mr, column = 2).value = "PREPARATION"
  
    #Comments
    print("Enter COMMENTS: " )
    x = input()
    sheet1.cell(row = mr, column = 3).value = x

    #Hours
    print("Enter LOCATION: ")
    x = input()
    sheet1.cell(row = mr, column = 4).value = x
    
    print("Enter TIME(mins): ")
    x = input()
    sheet1.cell(row = mr, column = 5).value = x

    wb1.save(home + '\\Desktop\\Book1.xlsx')
    count = count + 1
    print(str(count) + " line saved")
    print("Continue? : ")
    a = input()
    if a not in 'yYes':
        y = 2




